"""
Report Generator for CBSC Backtest System
Generates professional backtest reports in multiple formats
"""
import os
import io
import base64
from datetime import datetime
from typing import Dict, List, Optional, Any, Union
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import seaborn as sns
from jinja2 import Template, Environment, FileSystemLoader
import plotly.graph_objects as go
import plotly.express as px
from plotly.subplots import make_subplots

try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.graphics.shapes import Drawing
    from reportlab.graphics.charts.lineplots import LinePlot
    from reportlab.graphics.charts.barcharts import VerticalBarChart
    from reportlab.graphics.widgetbase import WidgetHolder
    from reportlab.graphics import renderPM
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import LineChart, Reference
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

from .enhanced_risk_metrics import EnhancedRiskMetrics


class ReportGenerator:
    """
    Generate professional backtest reports
    """

    def __init__(self, output_dir: str = "reports"):
        """
        Initialize ReportGenerator

        Args:
            output_dir: Directory to save reports
        """
        self.output_dir = output_dir
        self.ensure_output_dir()

        # Set up plotting style
        plt.style.use('seaborn-v0_8')
        sns.set_palette("husl")

        # Jinja2 environment for HTML reports
        self.jinja_env = Environment(
            loader=FileSystemLoader(os.path.join(os.path.dirname(__file__), 'templates'))
        )

    def ensure_output_dir(self):
        """Ensure output directory exists"""
        if not os.path.exists(self.output_dir):
            os.makedirs(self.output_dir)

    def generate_report(
        self,
        backtest_result: Dict,
        format: str = "pdf",
        include_charts: bool = True,
        template_name: str = "default"
    ) -> str:
        """
        Generate backtest report

        Args:
            backtest_result: Backtest results dictionary
            format: Output format ('pdf', 'excel', 'html', 'json')
            include_charts: Whether to include charts
            template_name: Template name for HTML reports

        Returns:
            str: Path to generated report
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        strategy_name = backtest_result.get('strategy_name', 'Unknown').replace(' ', '_')

        if format == "pdf":
            return self._generate_pdf_report(backtest_result, timestamp, include_charts)
        elif format == "excel":
            return self._generate_excel_report(backtest_result, timestamp, include_charts)
        elif format == "html":
            return self._generate_html_report(backtest_result, timestamp, template_name)
        elif format == "json":
            return self._generate_json_report(backtest_result, timestamp)
        else:
            raise ValueError(f"Unsupported format: {format}")

    def _generate_pdf_report(
        self,
        backtest_result: Dict,
        timestamp: str,
        include_charts: bool
    ) -> str:
        """Generate PDF report"""
        if not PDF_AVAILABLE:
            raise ImportError("ReportLab is required for PDF generation")

        filename = f"{self.output_dir}/backtest_report_{timestamp}.pdf"
        doc = SimpleDocTemplate(filename, pagesize=A4)

        # Get styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            spaceAfter=30,
            alignment=1  # Center
        )

        # Build story
        story = []

        # Title
        story.append(Paragraph("Backtest Performance Report", title_style))
        story.append(Spacer(1, 12))

        # Executive Summary
        story.append(Paragraph("Executive Summary", styles['Heading2']))
        summary_data = self._create_summary_table(backtest_result)
        story.append(summary_data)
        story.append(Spacer(1, 12))

        # Performance Metrics
        story.append(Paragraph("Performance Metrics", styles['Heading2']))
        metrics_data = self._create_metrics_table(backtest_result)
        story.append(metrics_data)
        story.append(Spacer(1, 12))

        # Risk Analysis
        story.append(Paragraph("Risk Analysis", styles['Heading2']))
        risk_data = self._create_risk_table(backtest_result)
        story.append(risk_data)
        story.append(Spacer(1, 12))

        # Trading Statistics
        story.append(Paragraph("Trading Statistics", styles['Heading2']))
        trading_data = self._create_trading_table(backtest_result)
        story.append(trading_data)
        story.append(Spacer(1, 12))

        # Add charts if requested
        if include_charts:
            charts = self._generate_report_charts(backtest_result)
            for chart_name, chart_path in charts.items():
                if chart_path and os.path.exists(chart_path):
                    story.append(Paragraph(chart_name.replace('_', ' ').title(), styles['Heading3']))
                    img = Image(chart_path, width=6*inch, height=4*inch)
                    story.append(img)
                    story.append(Spacer(1, 12))

        # Build PDF
        doc.build(story)

        # Clean up temporary chart files
        if include_charts:
            self._cleanup_chart_files(charts)

        return filename

    def _generate_excel_report(
        self,
        backtest_result: Dict,
        timestamp: str,
        include_charts: bool
    ) -> str:
        """Generate Excel report"""
        if not EXCEL_AVAILABLE:
            raise ImportError("OpenPyXL is required for Excel generation")

        filename = f"{self.output_dir}/backtest_report_{timestamp}.xlsx"
        wb = Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Summary Sheet
        self._create_summary_sheet(wb, backtest_result)

        # Performance Sheet
        self._create_performance_sheet(wb, backtest_result)

        # Risk Metrics Sheet
        self._create_risk_sheet(wb, backtest_result)

        # Trades Sheet
        self._create_trades_sheet(wb, backtest_result)

        # Monthly Returns Sheet
        self._create_monthly_returns_sheet(wb, backtest_result)

        # Save workbook
        wb.save(filename)

        return filename

    def _generate_html_report(
        self,
        backtest_result: Dict,
        timestamp: str,
        template_name: str
    ) -> str:
        """Generate HTML report"""
        filename = f"{self.output_dir}/backtest_report_{timestamp}.html"

        # Prepare template data
        template_data = self._prepare_template_data(backtest_result)

        # Generate charts as base64
        charts = self._generate_html_charts(backtest_result)
        template_data['charts'] = charts

        # Use default template if custom not found
        template_html = self._get_default_html_template()

        # Render template
        template = Template(template_html)
        html_content = template.render(**template_data)

        # Write to file
        with open(filename, 'w', encoding='utf-8') as f:
            f.write(html_content)

        return filename

    def _generate_json_report(self, backtest_result: Dict, timestamp: str) -> str:
        """Generate JSON report"""
        filename = f"{self.output_dir}/backtest_report_{timestamp}.json"

        # Prepare clean JSON data
        json_data = {
            'metadata': {
                'strategy_name': backtest_result.get('strategy_name'),
                'start_date': backtest_result.get('start_date'),
                'end_date': backtest_result.get('end_date'),
                'generated_at': datetime.now().isoformat()
            },
            'summary': backtest_result.get('summary', {}),
            'performance': backtest_result.get('performance_metrics', {}),
            'risk': backtest_result.get('risk_metrics', {}),
            'trading': backtest_result.get('trading_metrics', {}),
            'benchmark_comparison': backtest_result.get('benchmark_comparison', {}),
            'enhanced_metrics': {k: v for k, v in backtest_result.items()
                                if k not in ['strategy_name', 'start_date', 'end_date', 'summary',
                                           'performance_metrics', 'risk_metrics', 'trading_metrics',
                                           'benchmark_comparison', 'trades', 'portfolio_values', 'daily_returns']}
        }

        # Save as JSON
        import json
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(json_data, f, indent=2, default=str)

        return filename

    def _create_summary_table(self, backtest_result: Dict) -> Table:
        """Create summary table for PDF"""
        data = [
            ['Metric', 'Value'],
            ['Strategy Name', backtest_result.get('strategy_name', 'N/A')],
            ['Start Date', backtest_result.get('start_date', 'N/A')],
            ['End Date', backtest_result.get('end_date', 'N/A')],
            ['Initial Capital', f"${backtest_result.get('initial_capital', 0):,.2f}"],
            ['Final Capital', f"${backtest_result.get('final_capital', 0):,.2f}"],
            ['Total Return', f"{backtest_result.get('total_return', 0):.2%}"],
            ['Annual Return', f"{backtest_result.get('annualized_return', 0):.2%}"],
            ['Sharpe Ratio', f"{backtest_result.get('sharpe_ratio', 0):.3f}"],
            ['Max Drawdown', f"{backtest_result.get('max_drawdown', 0):.2%}"]
        ]

        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))

        return table

    def _create_metrics_table(self, backtest_result: Dict) -> Table:
        """Create performance metrics table"""
        metrics = backtest_result.get('metrics', {})

        data = [
            ['Performance Metric', 'Value'],
            ['Annual Volatility', f"{metrics.get('volatility', 0):.2%}"],
            ['Sortino Ratio', f"{metrics.get('sortino_ratio', 0):.3f}"],
            ['Calmar Ratio', f"{metrics.get('calmar_ratio', 0):.3f}"],
            ['Information Ratio', f"{metrics.get('information_ratio', 0):.3f}"],
            ['Total Trades', f"{metrics.get('total_trades', 0):,}"],
            ['Win Rate', f"{metrics.get('win_rate', 0):.2%}"],
            ['Profit Factor', f"{metrics.get('profit_factor', 0):.3f}"]
        ]

        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))

        return table

    def _create_risk_table(self, backtest_result: Dict) -> Table:
        """Create risk metrics table"""
        metrics = backtest_result.get('metrics', {})

        data = [
            ['Risk Metric', 'Value'],
            ['VaR (95%)', f"{metrics.get('var_95', 0):.2%}"],
            ['VaR (99%)', f"{metrics.get('var_99', 0):.2%}"],
            ['Expected Shortfall (95%)', f"{metrics.get('expected_shortfall_95', 0):.2%}"],
            ['Max Drawdown Duration', f"{metrics.get('max_drawdown_duration', 0)} days"],
            ['Omega Ratio', f"{metrics.get('omega_ratio', 0):.3f}"],
            ['Beta', f"{metrics.get('beta', 0):.3f}"],
            ['Alpha', f"{metrics.get('alpha', 0):.2%}"]
        ]

        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightcoral),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.mistyrose),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))

        return table

    def _create_trading_table(self, backtest_result: Dict) -> Table:
        """Create trading statistics table"""
        metrics = backtest_result.get('metrics', {})

        data = [
            ['Trading Statistic', 'Value'],
            ['Total Trades', f"{metrics.get('total_trades', 0):,}"],
            ['Winning Trades', f"{metrics.get('winning_trades', 0):,}"],
            ['Losing Trades', f"{metrics.get('losing_trades', 0):,}"],
            ['Average Win', f"${metrics.get('avg_win', 0):.2f}"],
            ['Average Loss', f"${metrics.get('avg_loss', 0):.2f}"],
            ['Total Commission', f"${metrics.get('total_commission', 0):,.2f}"],
            ['Total Slippage', f"${metrics.get('total_slippage', 0):,.2f}"]
        ]

        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgreen),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.lightgreen),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))

        return table

    def _generate_report_charts(self, backtest_result: Dict) -> Dict[str, str]:
        """Generate charts for report"""
        charts = {}

        try:
            # Portfolio Value Chart
            if 'portfolio_values' in backtest_result:
                charts['portfolio_value'] = self._create_portfolio_value_chart(
                    backtest_result['portfolio_values'],
                    backtest_result.get('daily_returns', [])
                )

            # Returns Distribution Chart
            if 'daily_returns' in backtest_result:
                charts['returns_distribution'] = self._create_returns_distribution_chart(
                    backtest_result['daily_returns']
                )

            # Drawdown Chart
            if 'portfolio_values' in backtest_result:
                charts['drawdown'] = self._create_drawdown_chart(
                    backtest_result['portfolio_values']
                )

            # Monthly Returns Heatmap
            if 'daily_returns' in backtest_result:
                charts['monthly_returns'] = self._create_monthly_returns_heatmap(
                    backtest_result['daily_returns']
                )

        except Exception as e:
            print(f"Warning: Error generating charts: {e}")

        return charts

    def _create_portfolio_value_chart(self, portfolio_values: List, returns: List) -> str:
        """Create portfolio value chart"""
        fig, ax = plt.subplots(figsize=(12, 6))

        # Convert to cumulative returns
        dates = pd.date_range(start='2023-01-01', periods=len(portfolio_values), freq='D')
        ax.plot(dates, portfolio_values, linewidth=2, label='Portfolio Value')

        # Add zero line
        ax.axhline(y=portfolio_values[0], color='gray', linestyle='--', alpha=0.7, label='Initial Value')

        ax.set_title('Portfolio Value Over Time', fontsize=16, fontweight='bold')
        ax.set_xlabel('Date')
        ax.set_ylabel('Portfolio Value ($)')
        ax.legend()
        ax.grid(True, alpha=0.3)

        # Format x-axis
        ax.xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m'))
        plt.xticks(rotation=45)

        plt.tight_layout()
        filename = f"{self.output_dir}/temp_portfolio_value.png"
        plt.savefig(filename, dpi=300, bbox_inches='tight')
        plt.close()

        return filename

    def _create_returns_distribution_chart(self, returns: List) -> str:
        """Create returns distribution chart"""
        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(15, 6))

        # Histogram
        ax1.hist(returns, bins=50, alpha=0.7, color='blue', edgecolor='black')
        ax1.set_title('Daily Returns Distribution', fontweight='bold')
        ax1.set_xlabel('Daily Return')
        ax1.set_ylabel('Frequency')
        ax1.grid(True, alpha=0.3)

        # QQ Plot
        from scipy import stats
        stats.probplot(returns, dist="norm", plot=ax2)
        ax2.set_title('Q-Q Plot', fontweight='bold')
        ax2.grid(True, alpha=0.3)

        plt.tight_layout()
        filename = f"{self.output_dir}/temp_returns_dist.png"
        plt.savefig(filename, dpi=300, bbox_inches='tight')
        plt.close()

        return filename

    def _create_drawdown_chart(self, portfolio_values: List) -> str:
        """Create drawdown chart"""
        values = np.array(portfolio_values)
        peak = np.maximum.accumulate(values)
        drawdown = (values - peak) / peak * 100

        fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(12, 10), gridspec_kw={'height_ratios': [3, 1]})

        dates = pd.date_range(start='2023-01-01', periods=len(portfolio_values), freq='D')

        # Portfolio value
        ax1.plot(dates, portfolio_values, color='blue', linewidth=2)
        ax1.set_title('Portfolio Value and Drawdown', fontsize=16, fontweight='bold')
        ax1.set_ylabel('Portfolio Value ($)')
        ax1.grid(True, alpha=0.3)

        # Drawdown
        ax2.fill_between(dates, drawdown, 0, color='red', alpha=0.3)
        ax2.plot(dates, drawdown, color='red', linewidth=1)
        ax2.set_ylabel('Drawdown (%)')
        ax2.set_xlabel('Date')
        ax2.grid(True, alpha=0.3)

        plt.tight_layout()
        filename = f"{self.output_dir}/temp_drawdown.png"
        plt.savefig(filename, dpi=300, bbox_inches='tight')
        plt.close()

        return filename

    def _create_monthly_returns_heatmap(self, returns: List) -> str:
        """Create monthly returns heatmap"""
        # Convert returns to DataFrame
        dates = pd.date_range(start='2023-01-01', periods=len(returns), freq='D')
        returns_series = pd.Series(returns, index=dates)

        # Group by month
        monthly_returns = returns_series.groupby([returns_series.index.year, returns_series.index.month]).apply(lambda x: (1 + x).prod() - 1)
        monthly_df = monthly_returns.unstack(fill_value=0)

        # Create heatmap
        plt.figure(figsize=(12, 8))
        sns.heatmap(
            monthly_df.T,
            annot=True,
            fmt='.2%',
            cmap='RdYlGn',
            center=0,
            cbar_kws={'label': 'Monthly Return'}
        )
        plt.title('Monthly Returns Heatmap', fontsize=16, fontweight='bold')
        plt.xlabel('Year')
        plt.ylabel('Month')

        filename = f"{self.output_dir}/temp_monthly_heatmap.png"
        plt.savefig(filename, dpi=300, bbox_inches='tight')
        plt.close()

        return filename

    def _generate_html_charts(self, backtest_result: Dict) -> Dict[str, str]:
        """Generate interactive charts for HTML report using Plotly"""
        charts = {}

        try:
            # Portfolio Value Chart
            if 'portfolio_values' in backtest_result:
                charts['portfolio_value'] = self._create_plotly_portfolio_chart(
                    backtest_result['portfolio_values']
                )

            # Returns Distribution
            if 'daily_returns' in backtest_result:
                charts['returns_dist'] = self._create_plotly_returns_dist(
                    backtest_result['daily_returns']
                )

        except Exception as e:
            print(f"Warning: Error generating Plotly charts: {e}")

        return charts

    def _create_plotly_portfolio_chart(self, portfolio_values: List) -> str:
        """Create interactive portfolio chart with Plotly"""
        dates = pd.date_range(start='2023-01-01', periods=len(portfolio_values), freq='D')

        fig = go.Figure()
        fig.add_trace(go.Scatter(
            x=dates,
            y=portfolio_values,
            mode='lines',
            name='Portfolio Value',
            line=dict(width=2, color='blue')
        ))

        fig.add_hline(
            y=portfolio_values[0],
            line_dash="dash",
            line_color="gray",
            annotation_text="Initial Value"
        )

        fig.update_layout(
            title='Portfolio Value Over Time',
            xaxis_title='Date',
            yaxis_title='Portfolio Value ($)',
            hovermode='x unified'
        )

        return fig.to_html(include_plotlyjs='cdn')

    def _create_plotly_returns_dist(self, returns: List) -> str:
        """Create interactive returns distribution with Plotly"""
        fig = make_subplots(
            rows=1, cols=2,
            subplot_titles=('Returns Distribution', 'Q-Q Plot'),
            specs=[[{"secondary_y": False}, {"secondary_y": False}]]
        )

        # Histogram
        fig.add_trace(
            go.Histogram(x=returns, nbinsx=50, name='Returns'),
            row=1, col=1
        )

        # QQ Plot
        from scipy import stats
        qq_data = stats.probplot(returns, dist="norm")
        fig.add_trace(
            go.Scatter(
                x=qq_data[0][0],
                y=qq_data[0][1],
                mode='markers',
                name='Data'
            ),
            row=1, col=2
        )
        fig.add_trace(
            go.Scatter(
                x=qq_data[0][0],
                y=qq_data[1][1] + qq_data[1][0] * qq_data[0][0],
                mode='lines',
                name='Normal'
            ),
            row=1, col=2
        )

        fig.update_layout(
            title='Returns Analysis',
            showlegend=False
        )

        return fig.to_html(include_plotlyjs='cdn')

    def _cleanup_chart_files(self, charts: Dict):
        """Clean up temporary chart files"""
        for chart_path in charts.values():
            if chart_path and os.path.exists(chart_path):
                os.remove(chart_path)

    def _prepare_template_data(self, backtest_result: Dict) -> Dict:
        """Prepare data for HTML template"""
        return {
            'title': f"Backtest Report - {backtest_result.get('strategy_name', 'Unknown')}",
            'generated_at': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            'summary': {
                'strategy_name': backtest_result.get('strategy_name'),
                'start_date': backtest_result.get('start_date'),
                'end_date': backtest_result.get('end_date'),
                'initial_capital': backtest_result.get('initial_capital', 0),
                'final_capital': backtest_result.get('final_capital', 0),
                'total_return': backtest_result.get('total_return', 0),
                'annualized_return': backtest_result.get('annualized_return', 0),
                'sharpe_ratio': backtest_result.get('sharpe_ratio', 0),
                'max_drawdown': backtest_result.get('max_drawdown', 0)
            },
            'metrics': backtest_result.get('metrics', {}),
            'trades': backtest_result.get('trades', [])
        }

    def _get_default_html_template(self) -> str:
        """Get default HTML template"""
        return """
<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
    <title>{{ title }}</title>
    <style>
        body { font-family: Arial, sans-serif; margin: 40px; background-color: #f5f5f5; }
        .container { max-width: 1200px; margin: 0 auto; background-color: white; padding: 30px; border-radius: 10px; box-shadow: 0 0 10px rgba(0,0,0,0.1); }
        h1 { color: #333; text-align: center; margin-bottom: 30px; }
        h2 { color: #666; border-bottom: 2px solid #ddd; padding-bottom: 10px; }
        .summary { display: grid; grid-template-columns: repeat(auto-fit, minmax(250px, 1fr)); gap: 20px; margin: 20px 0; }
        .metric-card { background-color: #f8f9fa; padding: 20px; border-radius: 8px; text-align: center; }
        .metric-value { font-size: 2em; font-weight: bold; color: #007bff; }
        .metric-label { color: #666; margin-top: 5px; }
        table { width: 100%; border-collapse: collapse; margin: 20px 0; }
        th, td { padding: 12px; text-align: left; border-bottom: 1px solid #ddd; }
        th { background-color: #f8f9fa; font-weight: bold; }
        .positive { color: green; }
        .negative { color: red; }
        .chart-container { margin: 30px 0; }
        footer { text-align: center; color: #999; margin-top: 50px; font-size: 0.9em; }
    </style>
</head>
<body>
    <div class="container">
        <h1>{{ title }}</h1>
        <p style="text-align: center; color: #666;">Generated on: {{ generated_at }}</p>

        <h2>Executive Summary</h2>
        <div class="summary">
            <div class="metric-card">
                <div class="metric-value">${{ "%.2f"|format(summary.initial_capital) }}</div>
                <div class="metric-label">Initial Capital</div>
            </div>
            <div class="metric-card">
                <div class="metric-value">${{ "%.2f"|format(summary.final_capital) }}</div>
                <div class="metric-label">Final Capital</div>
            </div>
            <div class="metric-card">
                <div class="metric-value {{ 'positive' if summary.total_return > 0 else 'negative' }}">
                    {{ "%.2f%%"|format(summary.total_return * 100) }}
                </div>
                <div class="metric-label">Total Return</div>
            </div>
            <div class="metric-card">
                <div class="metric-value {{ 'positive' if summary.sharpe_ratio > 1 else 'negative' }}">
                    {{ "%.3f"|format(summary.sharpe_ratio) }}
                </div>
                <div class="metric-label">Sharpe Ratio</div>
            </div>
        </div>

        <h2>Performance Metrics</h2>
        <table>
            <tr><th>Metric</th><th>Value</th></tr>
            <tr><td>Annual Return</td><td>{{ "%.2f%%"|format(summary.annualized_return * 100) }}</td></tr>
            <tr><td>Max Drawdown</td><td class="negative">{{ "%.2f%%"|format(summary.max_drawdown * 100) }}</td></tr>
            <tr><td>Total Trades</td><td>{{ metrics.total_trades if metrics.total_trades else 0 }}</td></tr>
            <tr><td>Win Rate</td><td>{{ "%.2f%%"|format(metrics.win_rate * 100) if metrics.win_rate else "N/A" }}</td></tr>
        </table>

        {% if charts %}
            {% for chart_name, chart_html in charts.items() %}
            <div class="chart-container">
                <h3>{{ chart_name|replace('_', ' ')|title }}</h3>
                {{ chart_html|safe }}
            </div>
            {% endfor %}
        {% endif %}

        <footer>
            <p>Report generated by CBSC Backtest System</p>
        </footer>
    </div>
</body>
</html>
        """

    def _create_summary_sheet(self, wb: Workbook, backtest_result: Dict):
        """Create summary sheet in Excel"""
        ws = wb.create_sheet("Summary", 0)

        # Title
        ws['A1'] = "Backtest Performance Report"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:B1')

        # Summary data
        summary_data = [
            ['Metric', 'Value'],
            ['Strategy Name', backtest_result.get('strategy_name', 'N/A')],
            ['Start Date', backtest_result.get('start_date', 'N/A')],
            ['End Date', backtest_result.get('end_date', 'N/A')],
            ['Initial Capital', backtest_result.get('initial_capital', 0)],
            ['Final Capital', backtest_result.get('final_capital', 0)],
            ['Total Return', backtest_result.get('total_return', 0)],
            ['Annual Return', backtest_result.get('annualized_return', 0)],
            ['Sharpe Ratio', backtest_result.get('sharpe_ratio', 0)],
            ['Max Drawdown', backtest_result.get('max_drawdown', 0)]
        ]

        for row_idx, row_data in enumerate(summary_data, start=3):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)

                # Header styling
                if row_idx == 3:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20

    def _create_performance_sheet(self, wb: Workbook, backtest_result: Dict):
        """Create performance metrics sheet"""
        ws = wb.create_sheet("Performance")

        metrics = backtest_result.get('metrics', {})

        performance_data = [
            ['Performance Metric', 'Value'],
            ['Annual Volatility', metrics.get('volatility', 0)],
            ['Sortino Ratio', metrics.get('sortino_ratio', 0)],
            ['Calmar Ratio', metrics.get('calmar_ratio', 0)],
            ['Information Ratio', metrics.get('information_ratio', 0)],
            ['Total Trades', metrics.get('total_trades', 0)],
            ['Winning Trades', metrics.get('winning_trades', 0)],
            ['Losing Trades', metrics.get('losing_trades', 0)],
            ['Win Rate', metrics.get('win_rate', 0)],
            ['Profit Factor', metrics.get('profit_factor', 0)],
            ['Average Win', metrics.get('avg_win', 0)],
            ['Average Loss', metrics.get('avg_loss', 0)],
            ['Total Commission', metrics.get('total_commission', 0)],
            ['Total Slippage', metrics.get('total_slippage', 0)]
        ]

        for row_idx, row_data in enumerate(performance_data, start=1):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)

                # Header styling
                if row_idx == 1:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")

                # Format percentage values
                if row_idx > 1 and isinstance(value, (int, float)):
                    if 'Rate' in performance_data[row_idx-1][0] or performance_data[row_idx-1][0] == 'Annual Volatility':
                        cell.number_format = '0.00%'

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20

    def _create_risk_sheet(self, wb: Workbook, backtest_result: Dict):
        """Create risk metrics sheet"""
        ws = wb.create_sheet("Risk Metrics")

        metrics = backtest_result.get('metrics', {})

        risk_data = [
            ['Risk Metric', 'Value'],
            ['VaR (95%)', metrics.get('var_95', 0)],
            ['VaR (99%)', metrics.get('var_99', 0)],
            ['Expected Shortfall (95%)', metrics.get('expected_shortfall_95', 0)],
            ['Max Drawdown', metrics.get('max_drawdown', 0)],
            ['Max Drawdown Duration', metrics.get('max_drawdown_duration', 0)],
            ['Omega Ratio', metrics.get('omega_ratio', 0)],
            ['Beta', metrics.get('beta', 0)],
            ['Alpha', metrics.get('alpha', 0)],
            ['Treynor Ratio', metrics.get('treynor_ratio', 0)],
            ['Jensen Alpha', metrics.get('jensen_alpha', 0)],
            ['Relative Sharpe Ratio', metrics.get('relative_sharpe_ratio', 0)],
            ['Upside Capture', metrics.get('upside_capture', 0)],
            ['Downside Capture', metrics.get('downside_capture', 0)]
        ]

        for row_idx, row_data in enumerate(risk_data, start=1):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)

                # Header styling
                if row_idx == 1:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")

                # Format percentage values
                if row_idx > 1 and isinstance(value, (int, float)):
                    if any(keyword in risk_data[row_idx-1][0] for keyword in ['VaR', 'Shortfall', 'Drawdown', 'Capture']):
                        cell.number_format = '0.00%'

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20

    def _create_trades_sheet(self, wb: Workbook, backtest_result: Dict):
        """Create trades detail sheet"""
        ws = wb.create_sheet("Trades")

        trades = backtest_result.get('trades', [])

        if not trades:
            ws['A1'] = "No trades data available"
            return

        # Headers
        headers = ['Date', 'Symbol', 'Action', 'Quantity', 'Price', 'Commission', 'Slippage', 'PnL']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

        # Trade data
        for row_idx, trade in enumerate(trades, start=2):
            if isinstance(trade, dict):
                ws.cell(row=row_idx, column=1, value=trade.get('date', 'N/A'))
                ws.cell(row=row_idx, column=2, value=trade.get('symbol', 'N/A'))
                ws.cell(row=row_idx, column=3, value=trade.get('action', 'N/A'))
                ws.cell(row=row_idx, column=4, value=trade.get('quantity', 0))
                ws.cell(row=row_idx, column=5, value=trade.get('price', 0))
                ws.cell(row=row_idx, column=6, value=trade.get('commission', 0))
                ws.cell(row=row_idx, column=7, value=trade.get('slippage', 0))
                ws.cell(row=row_idx, column=8, value=trade.get('pnl', 0))

        # Auto-adjust columns
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col_idx)].width = 15

    def _create_monthly_returns_sheet(self, wb: Workbook, backtest_result: Dict):
        """Create monthly returns sheet"""
        ws = wb.create_sheet("Monthly Returns")

        daily_returns = backtest_result.get('daily_returns', [])
        if not daily_returns:
            ws['A1'] = "No returns data available"
            return

        # Create date index
        dates = pd.date_range(start='2023-01-01', periods=len(daily_returns), freq='D')
        returns_series = pd.Series(daily_returns, index=dates)

        # Group by month and year
        monthly_returns = returns_series.groupby([returns_series.index.year, returns_series.index.month]).apply(
            lambda x: (1 + x).prod() - 1
        )
        monthly_df = monthly_returns.unstack(fill_value=0)

        # Write headers
        ws.cell(row=1, column=1, value="Year / Month")
        for col_idx, month in enumerate(monthly_df.columns, start=2):
            month_name = pd.to_datetime(f"2023-{month:02d}").strftime('%B')
            ws.cell(row=1, column=col_idx, value=month_name)
            ws.cell(row=1, column=col_idx).font = Font(bold=True)
            ws.cell(row=1, column=col_idx).fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

        # Write data
        for row_idx, (year, row_data) in enumerate(monthly_df.iterrows(), start=2):
            ws.cell(row=row_idx, column=1, value=year)
            ws.cell(row=row_idx, column=1).font = Font(bold=True)

            for col_idx, value in enumerate(row_data, start=2):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.number_format = '0.00%'

                # Color coding
                if value > 0:
                    cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                elif value < 0:
                    cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")

        # Add total row
        total_row = len(monthly_df) + 3
        ws.cell(row=total_row, column=1, value="Average")
        ws.cell(row=total_row, column=1).font = Font(bold=True)

        for col_idx in range(2, len(monthly_df.columns) + 2):
            avg = monthly_df.iloc[:, col_idx-2].mean()
            cell = ws.cell(row=total_row, column=col_idx, value=avg)
            cell.number_format = '0.00%'
            cell.font = Font(bold=True)