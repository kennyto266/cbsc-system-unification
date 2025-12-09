#!/usr / bin / env python3
"""
Excel格式導出器
支持多工作表、圖表、條件格式化等高級功能
"""

import logging
from pathlib import Path
from typing import Any, Dict, List

import numpy as np
import pandas as pd

logger = logging.getLogger(__name__)


class ExcelExporter:
    """Excel導出器類"""

    def __init__(self, config: Dict):
        self.config = config
        self.engine = config.get("engine", "openpyxl")
        self.include_formulas = config.get("include_formulas", True)
        self.conditional_formatting = config.get("conditional_formatting", True)
        self.auto_column_width = config.get("auto_column_width", True)
        self.max_rows_per_sheet = config.get("max_rows_per_sheet", 100000)

    def export(self, data: Any, output_path: Path, options: Dict = None) -> bool:
        """
        導出數據到Excel文件

        Args:
            data: 要導出的數據
            output_path: 輸出文件路徑
            options: 導出選項

        Returns:
            bool: 是否成功
        """
        try:
            # 創建Excel writer
            with pd.ExcelWriter(output_path, engine = self.engine, mode="w") as writer:

                # 根據數據類型進行不同的處理
                if isinstance(data, dict):
                    self._export_dict_data(writer, data, options)
                elif isinstance(data, pd.DataFrame):
                    self._export_dataframe(writer, data, options)
                elif isinstance(data, list):
                    self._export_list_data(writer, data, options)
                else:
                    # 轉換為DataFrame後導出
                    df = pd.DataFrame([data])
                    df.to_excel(writer, sheet_name="Data", index = False)

                # 添加圖表（如果配置啟用）
                if options and options.get("include_charts", True):
                    self._add_charts(writer, data, output_path)

            logger.info(f"✅ Excel文件導出成功: {output_path}")
            return True

        except Exception as e:
            logger.error(f"❌ Excel導出失敗: {e}")
            return False

    def _export_dict_data(
        self, writer: pd.ExcelWriter, data: Dict, options: Dict = None
    ):
        """導出字典數據到多個工作表"""
        sheet_index = 0

        for key, value in data.items():
            # 跳過元數據
            if key.lower() in ["metadata", "config", "settings"]:
                continue

            try:
                if isinstance(value, pd.DataFrame):
                    # 導出DataFrame
                    self._export_dataframe(
                        writer, value, options, sheet_name = str(key)[:31]
                    )
                elif isinstance(value, (dict, list)):
                    # 轉換為DataFrame
                    df = (
                        pd.DataFrame(value)
                        if isinstance(value, list)
                        else pd.DataFrame([value])
                    )
                    self._export_dataframe(
                        writer, df, options, sheet_name = str(key)[:31]
                    )
                else:
                    # 單個值
                    df = pd.DataFrame({"Value": [value]})
                    df.to_excel(writer, sheet_name = str(key)[:31], index = False)

                sheet_index += 1

            except Exception as e:
                logger.warning(f"⚠️ 工作表 {key} 導出失敗: {e}")

    def _export_dataframe(
        self,
        writer: pd.ExcelWriter,
        df: pd.DataFrame,
        options: Dict = None,
        sheet_name: str = None,
    ):
        """導出DataFrame到Excel工作表"""
        if df.empty:
            return

        # 處理大數據集
        if len(df) > self.max_rows_per_sheet:
            self._export_large_dataframe(writer, df, options, sheet_name)
            return

        if sheet_name is None:
            sheet_name = "Data"

        try:
            # 導出DataFrame
            df.to_excel(writer, sheet_name = sheet_name, index = False)

            # 獲取工作簿和工作表對象
            if self.engine == "openpyxl":
                from openpyxl import load_workbook
                from openpyxl.chart import BarChart, LineChart, Reference
                from openpyxl.styles import Alignment, Font, PatternFill
                from openpyxl.utils.dataframe import dataframe_to_rows

                # 獲取當前工作表
                worksheet = writer.sheets[sheet_name]

                # 自動調整列寬
                if self.auto_column_width:
                    self._auto_adjust_column_width(worksheet, df)

                # 添加條件格式化
                if self.conditional_formatting:
                    self._add_conditional_formatting(worksheet, df)

                # 添加標題格式
                self._format_header(worksheet, df)

        except Exception as e:
            logger.warning(f"⚠️ Excel格式化失敗: {e}")

    def _export_large_dataframe(
        self,
        writer: pd.ExcelWriter,
        df: pd.DataFrame,
        options: Dict = None,
        sheet_name: str = None,
    ):
        """導出大DataFrame到多個工作表"""
        if sheet_name is None:
            sheet_name = "Data"

        max_rows = self.max_rows_per_sheet - 10  # 預留標題和格式空間
        total_rows = len(df)
        num_sheets = (total_rows // max_rows) + 1

        for i in range(num_sheets):
            start_idx = i * max_rows
            end_idx = min((i + 1) * max_rows, total_rows)
            df_subset = df.iloc[start_idx:end_idx]

            sheet_name_with_index = f"{sheet_name}_{i + 1}"
            if len(sheet_name_with_index) > 31:
                sheet_name_with_index = f"Sheet_{i + 1}"

            df_subset.to_excel(writer, sheet_name = sheet_name_with_index, index = False)

    def _auto_adjust_column_width(self, worksheet, df: pd.DataFrame):
        """自動調整列寬"""
        try:
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter

                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception:
                        pass

                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width

        except Exception as e:
            logger.debug(f"自動調整列寬失敗: {e}")

    def _add_conditional_formatting(self, worksheet, df: pd.DataFrame):
        """添加條件格式化"""
        try:
            from openpyxl.formatting.rule import CellIsRule
            from openpyxl.styles import PatternFill

            # 為數值列添加條件格式
            numeric_columns = df.select_dtypes(include=[np.number]).columns

            for idx, col_name in enumerate(numeric_columns):
                col_letter = chr(65 + idx)  # A, B, C, ...

                # 正數綠色，負數紅色
                green_fill = PatternFill(
                    start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
                )
                red_fill = PatternFill(
                    start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
                )

                positive_rule = CellIsRule(
                    operator="greaterThan", formula=["0"], fill = green_fill
                )
                negative_rule = CellIsRule(
                    operator="lessThan", formula=["0"], fill = red_fill
                )

                worksheet.conditional_formatting.add(
                    f"{col_letter}2:{col_letter}{len(df)+1}", positive_rule
                )
                worksheet.conditional_formatting.add(
                    f"{col_letter}2:{col_letter}{len(df)+1}", negative_rule
                )

        except Exception as e:
            logger.debug(f"條件格式化添加失敗: {e}")

    def _format_header(self, worksheet, df: pd.DataFrame):
        """格式化標題行"""
        try:
            from openpyxl.styles import Alignment, Font, PatternFill

            # 獲取標題行
            header_row = worksheet[1]

            # 設置標題樣式
            header_font = Font(bold = True, color="FFFFFF")
            header_fill = PatternFill(
                start_color="4472C4", end_color="4472C4", fill_type="solid"
            )
            header_alignment = Alignment(horizontal="center", vertical="center")

            for cell in header_row:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # 凍結首行
            worksheet.freeze_panes = "A2"

        except Exception as e:
            logger.debug(f"標題格式化失敗: {e}")

    def _add_charts(self, writer: pd.ExcelWriter, data: Any, output_path: Path):
        """添加圖表到Excel文件"""
        try:
            if self.engine != "openpyxl":
                return

            # 這裡可以添加圖表生成邏輯
            # 由於圖表生成比較複雜，暫時跳過

        except Exception as e:
            logger.debug(f"圖表添加失敗: {e}")

    def _export_list_data(
        self, writer: pd.ExcelWriter, data: List, options: Dict = None
    ):
        """導出列表數據"""
        try:
            if len(data) == 0:
                return

            # 如果列表元素是字典，轉換為DataFrame
            if isinstance(data[0], dict):
                df = pd.DataFrame(data)
                self._export_dataframe(writer, df, options, sheet_name="List_Data")
            else:
                # 簡單列表
                df = pd.DataFrame({"Value": data})
                self._export_dataframe(writer, df, options, sheet_name="List_Data")

        except Exception as e:
            logger.error(f"列表數據導出失敗: {e}")

    def export_summary_report(
        self, summary_data: Dict[str, Any], output_path: Path
    ) -> bool:
        """
        導出摘要報告到Excel

        Args:
            summary_data: 摘要數據
            output_path: 輸出路徑

        Returns:
            bool: 是否成功
        """
        try:
            with pd.ExcelWriter(output_path, engine = self.engine) as writer:

                # 摘要表
                summary_df = pd.DataFrame(
                    [
                        {
                            "指標": "總回報率",
                            "數值": f"{summary_data.get('total_return', 0):.2%}",
                        },
                        {
                            "指標": "年化回報率",
                            "數值": f"{summary_data.get('annual_return', 0):.2%}",
                        },
                        {
                            "指標": "Sharpe比率",
                            "數值": f"{summary_data.get('sharpe_ratio', 0):.3f}",
                        },
                        {
                            "指標": "最大回撤",
                            "數值": f"{summary_data.get('max_drawdown', 0):.2%}",
                        },
                        {
                            "指標": "波動率",
                            "數值": f"{summary_data.get('volatility', 0):.2%}",
                        },
                        {
                            "指標": "交易次數",
                            "數值": summary_data.get("trade_count", 0),
                        },
                        {
                            "指標": "勝率",
                            "數值": f"{summary_data.get('win_rate', 0):.2%}",
                        },
                    ]
                )

                summary_df.to_excel(writer, sheet_name="策略摘要", index = False)

                # 如果有交易記錄，導出交易詳情
                if "trades" in summary_data and summary_data["trades"]:
                    trades_df = pd.DataFrame(summary_data["trades"])
                    trades_df.to_excel(writer, sheet_name="交易記錄", index = False)

                # 如果有每日業績，導出業績數據
                if "daily_returns" in summary_data:
                    daily_df = pd.DataFrame(summary_data["daily_returns"])
                    daily_df.to_excel(writer, sheet_name="每日業績", index = False)

            logger.info(f"✅ 摘要報告導出成功: {output_path}")
            return True

        except Exception as e:
            logger.error(f"❌ 摘要報告導出失敗: {e}")
            return False
